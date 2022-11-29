import win32com.client

import openpyxl




def DisableLinks(filepath):
    wrkbk = openpyxl.load_workbook(filepath)
    sh = wrkbk.active
    logs = []
    for i in range(1,sh.max_row + 1):
        for j in range(1,sh.max_column+1):
            print(sh.cell(row=i,column=j)._hyperlink)
            if sh.cell(row=i,column=j).hyperlink is None:
                continue
            log = f"[-]Remove the Links ( {str(sh.cell(row=i,column=j).hyperlink)} )"  
            print(log)
            logs.append(log)                          
            sh.cell(row=i,column=j).hyperlink = ''
    wrkbk.close()
    wrkbk.save(filepath)
    del sh,wrkbk
    return logs


def DisableMacros(path_to_file):
    # open excel app and the workbook
    xlApp = win32com.client.Dispatch("Excel.Application")
    xlwb = xlApp.Workbooks.Open(path_to_file)
    logs = []
    # ITERATE THROUGH EACH VB COMPONENT (CLASS MODULE, STANDARD MODULE, USER FORMS)
    try:    
        for i in xlwb.VBProject.VBComponents:     
            log = f'[-]Remove Macro =>> {i.Name}'
            print(log)
            logs.append(log)   
            xlmodule = xlwb.VBProject.VBComponents(i.Name)
            if xlmodule.Type in [1, 2, 3]:            
                xlwb.VBProject.VBComponents.Remove(xlmodule)

    except Exception as e:
        print(e)
    finally:    
        # CLOSE AND SAVE AND UNINITIALIZE APP
        print('==================Successfully Removed Marcos From excel File=====================')
        xlwb.Close(True)
        xlApp.Quit
        xlApp = None
        return logs

DisableMacros("zdoni.xlsm")